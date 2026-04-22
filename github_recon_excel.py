#!/usr/bin/env python3
"""
GitHub Recon - Excel Report Generator

Usage (CLI):
    python3 github_recon_excel.py "pentest agent" --limit 50 -o report.xlsx

Usage (from Rust):
    python3 github_recon_excel.py --json /tmp/repos.json --output report.xlsx "pentest agent"

Supports any query - categories auto-detected from query keywords.
"""

import json, os, sys, urllib.request, urllib.parse, argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

GH_TOKEN = os.environ.get("GITHUB_TOKEN", "")

def search_github(query, limit=100):
    url = f"https://api.github.com/search/repositories?q={urllib.parse.quote(query)}&per_page={min(limit,100)}&sort=stars&order=desc"
    req = urllib.request.Request(url)
    if GH_TOKEN:
        req.add_header("Authorization", f"Bearer {GH_TOKEN}")
    req.add_header("User-Agent", "github-recon/1.0")
    with urllib.request.urlopen(req, timeout=30) as r:
        return json.loads(r.read())

def detect_category(query):
    q = query.lower()
    cats = [
        ("pentest", "Pentest"), ("penetration testing", "Pentest"),
        ("vulnerability", "Vulnerability"), ("cve", "Vulnerability"),
        ("bug bounty", "Bug Bounty"), ("bounty", "Bug Bounty"),
        ("recon", "Reconnaissance"), ("reconnaissance", "Reconnaissance"),
        ("network", "Network Discovery"), ("port scan", "Network Discovery"),
        ("host discovery", "Network Discovery"), ("scanner", "Scanner"),
        ("mcp", "MCP"), ("ai agent", "AI Agent"), ("llm agent", "AI Agent"),
        ("autonomous", "AI Agent"), ("agentic", "AI Agent"),
        ("automation", "Automation"), ("exploit", "Exploit"),
        ("security", "Security"),
    ]
    for kw, cat in cats:
        if kw in q:
            return cat
    return None

def write_xlsx(repos, path, query):
    wb = Workbook()
    fnt = Font(bold=True, color="FFFFFF", size=11)
    fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdrs = ["#", "Repository", "URL", "Description", "Stars", "Forks", "Language", "Updated"]
    ws = wb.active
    ws.title = "Results"
    for col, w in zip('ABCDEFGH', [5, 40, 55, 65, 8, 8, 12, 15]):
        ws.column_dimensions[col].width = w
    ws['A1'] = f"Query: {query}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Total: {len(repos)} repositories"
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = fnt; c.fill = fill
    sorted_repos = sorted(repos, key=lambda x: x['stargazers_count'], reverse=True)
    for i, r in enumerate(sorted_repos):
        row = i + 5
        ws.cell(row, 1, i+1); ws.cell(row, 2, r['full_name']); ws.cell(row, 3, r['html_url'])
        ws.cell(row, 4, r.get('description') or ''); ws.cell(row, 5, r['stargazers_count'])
        ws.cell(row, 6, r['forks_count']); ws.cell(row, 7, r.get('language') or '?'); ws.cell(row, 8, r['updated_at'][:10])
    cat = detect_category(query)
    if cat:
        ws2 = wb.create_sheet(title=cat[:31])
        for col, w in zip('ABCDEFGH', [5, 40, 55, 65, 8, 8, 12, 15]):
            ws2.column_dimensions[col].width = w
        ws2['A1'] = f"Category: {cat}"; ws2['A1'].font = fnt; ws2['A1'].fill = fill
        ws2['A2'] = f"Total: {len(repos)} repos"
        for col, h in enumerate(hdrs, 1):
            c = ws2.cell(row=3, column=col, value=h)
            c.font = fnt; c.fill = fill
        for i, r in enumerate(sorted_repos):
            row = i + 4
            ws2.cell(row, 1, i+1); ws2.cell(row, 2, r['full_name']); ws2.cell(row, 3, r['html_url'])
            ws2.cell(row, 4, r.get('description') or ''); ws2.cell(row, 5, r['stargazers_count'])
            ws2.cell(row, 6, r['forks_count']); ws2.cell(row, 7, r.get('language') or '?'); ws2.cell(row, 8, r['updated_at'][:10])
        ws2.freeze_panes = 'A4'
    ws3 = wb.create_sheet(title="All Combined")
    for col, w in zip('ABCDEFGH', [5, 40, 55, 65, 8, 8, 12, 15]):
        ws3.column_dimensions[col].width = w
    ws3['A1'] = "All Repositories Combined"; ws3['A1'].font = fnt; ws3['A1'].fill = fill
    ws3['A2'] = f"Total: {len(repos)} repositories"
    for col, h in enumerate(hdrs, 1):
        c = ws3.cell(row=3, column=col, value=h)
        c.font = fnt; c.fill = fill
    for i, r in enumerate(sorted_repos):
        row = i + 4
        ws3.cell(row, 1, i+1); ws3.cell(row, 2, r['full_name']); ws3.cell(row, 3, r['html_url'])
        ws3.cell(row, 4, r.get('description') or ''); ws3.cell(row, 5, r['stargazers_count'])
        ws3.cell(row, 6, r['forks_count']); ws3.cell(row, 7, r.get('language') or '?'); ws3.cell(row, 8, r['updated_at'][:10])
    ws3.freeze_panes = 'A4'
    wb.save(path)
    return len(repos)

def main():
    parser = argparse.ArgumentParser(description="GitHub Recon - Excel Report Generator")
    parser.add_argument("query", nargs="?", default="", help="Search query string")
    parser.add_argument("-o", "--output", help="Output file path")
    parser.add_argument("-l", "--limit", type=int, default=100, help="Max results")
    parser.add_argument("--json", dest="json_file", help="JSON file with repos (instead of searching)")
    parser.add_argument("--format", choices=["xlsx", "csv", "both"], default="xlsx")
    args = parser.parse_args()

    if args.json_file:
        # Called from Rust with pre-fetched repos
        with open(args.json_file) as f:
            repos = json.load(f)
        query = args.query or "Report"
        out_path = args.output or "report.xlsx"
    else:
        if not args.query:
            parser.print_help()
            return
        print(f"Searching GitHub for: {args.query}")
        data = search_github(args.query, args.limit)
        repos = data.get('items', [])
        print(f"Found {len(repos)} repositories")
        out_path = args.output or f"{args.query[:30].replace(' ', '_')}_report.xlsx"

    if not repos:
        print("No repositories found.")
        return

    count = write_xlsx(repos, out_path, args.query or "Report")
    print(f"Wrote {count} repos to: {out_path}")
    if os.environ.get("VERBOSE"):
        print(f"Size: {os.path.getsize(out_path)/1024:.1f} KB")

if __name__ == "__main__":
    main()