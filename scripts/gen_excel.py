#!/usr/bin/env python3
"""GitHub Recon - Excel generator (called from Rust CLI)"""
import csv, sys, os, argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

parser = argparse.ArgumentParser()
parser.add_argument('--csv', required=True)
parser.add_argument('--output', required=True)
parser.add_argument('--query', required=True)
args = parser.parse_args()

repos = []
with open(args.csv) as f:
    reader = csv.DictReader(f)
    for row in reader:
        repos.append(row)

wb = Workbook()
fnt = Font(bold=True, color='FFFFFF', size=11)
fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
hdrs = ['#', 'Repository', 'URL', 'Description', 'Stars', 'Forks', 'Language', 'Updated']

ws = wb.active
ws.title = 'Results'
for col, w in zip('ABCDEFGH', [5, 40, 55, 65, 8, 8, 12, 15]):
    ws.column_dimensions[col].width = w
ws['A1'] = 'Query: ' + args.query
ws['A1'].font = Font(bold=True, size=14)
ws['A2'] = 'Total: ' + str(len(repos)) + ' repositories'
for col, h in enumerate(hdrs, 1):
    c = ws.cell(row=4, column=col, value=h)
    c.font = fnt
    c.fill = fill

sorted_repos = sorted(repos, key=lambda x: int(x.get('stargazers_count', 0)), reverse=True)
for i, r in enumerate(sorted_repos):
    row = i + 5
    ws.cell(row, 1, i+1)
    ws.cell(row, 2, r['full_name'])
    ws.cell(row, 3, r['html_url'])
    ws.cell(row, 4, r.get('description', ''))
    ws.cell(row, 5, int(r.get('stargazers_count', 0)))
    ws.cell(row, 6, int(r.get('forks_count', 0)))
    ws.cell(row, 7, r.get('language', '?'))
    ws.cell(row, 8, r.get('updated_at', '')[:10])

ws3 = wb.create_sheet(title='All Combined')
for col, w in zip('ABCDEFGH', [5, 40, 55, 65, 8, 8, 12, 15]):
    ws3.column_dimensions[col].width = w
ws3['A1'] = 'All Repositories Combined'
ws3['A1'].font = fnt
ws3['A1'].fill = fill
ws3['A2'] = 'Total: ' + str(len(repos)) + ' repositories'
for col, h in enumerate(hdrs, 1):
    c = ws3.cell(row=3, column=col, value=h)
    c.font = fnt
    c.fill = fill
for i, r in enumerate(sorted_repos):
    row = i + 4
    ws3.cell(row, 1, i+1)
    ws3.cell(row, 2, r['full_name'])
    ws3.cell(row, 3, r['html_url'])
    ws3.cell(row, 4, r.get('description', ''))
    ws3.cell(row, 5, int(r.get('stargazers_count', 0)))
    ws3.cell(row, 6, int(r.get('forks_count', 0)))
    ws3.cell(row, 7, r.get('language', '?'))
    ws3.cell(row, 8, r.get('updated_at', '')[:10])
ws3.freeze_panes = 'A4'

wb.save(args.output)
os.unlink(args.csv)
print('Generated ' + args.output + ' with ' + str(len(repos)) + ' repos')